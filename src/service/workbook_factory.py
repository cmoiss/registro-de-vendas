from openpyxl import Workbook, load_workbook 

class WorkBookFactory:
    def __init__(self):
        self.workboook_name = "Planilha de vendas.xlsx"
    
    def _gerar_sheet_vendas(self, book: Workbook):
        active_sheet = book.active
        active_sheet.title = "Vendas"
        active_sheet.append(["Nome do Produto", "Valor", "Nome do Cliente", "Forma de Pagamento", "Data da Venda", "Horário da Venda", "Vendedor"])


    def _gerar_sheet_produtos(self, book: Workbook):
        book.create_sheet("Produtos")
        sheet_produtos = book["Produtos"]
        sheet_produtos.append(["Nome do Produto", "Valor"])
        
    def _generate_template_workbook(self):
        book = Workbook()
        self._gerar_sheet_vendas(book)
        self._gerar_sheet_produtos(book)
        book.save(self.workboook_name)
        return book

    def load_book(self):
        try:
            book = load_workbook(self.workboook_name)
            return book
        except FileNotFoundError:
            return self._generate_template_workbook()
        
WorkBookFactory().load_book()