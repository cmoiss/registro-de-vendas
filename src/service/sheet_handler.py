from openpyxl import load_workbook
from service.workbook_factory import WorkBookFactory
from model.venda import Venda
from model.produto import Produto

class SheetHandler:
    def __init__(self):
        self.nome_planilha = WorkBookFactory().workboook_name
    
    def registrar_venda(self, venda: Venda):
        factoy = WorkBookFactory()
        book = factoy.load_book()

        sheet_produtos = book["Vendas"]
        sheet_produtos.append([
            venda.produto.nome,
            venda.produto.preco,
            venda.cliente.name, 
            venda.forma_pagamento.forma,
            None,
            None,
            # venda.data_hora_venda,
            venda.vendedor.vendedor
        ])
        
        book.save(self.nome_planilha)
    
    def load_products(self):
        # 1. Carregar a planilha
        planilha = load_workbook(self.nome_planilha)
        sheet_produtos = planilha["Produtos"]

        # 2. Extrair dados e criar objetos Produto
        produtos_list = []
        for linha in sheet_produtos.iter_rows(min_row=2, values_only=True):  # Pula cabeçalho
            nome = linha[0]
            valor = linha[1]
            produto = Produto(nome=nome, preco=valor)  # Criando objeto Produto
            produtos_list.append(produto)  # Adicionando o objeto à lista

        return produtos_list